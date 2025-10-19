#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Batch extract 8 core procurement indicators from HTML announcements,
and export the results to CSV / JSONL / XLSX files.
"""
from __future__ import annotations

import argparse
import csv
import json
import logging
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

from bs4 import BeautifulSoup
from bs4.element import NavigableString, Tag
from dateutil import parser as date_parser
from openpyxl import Workbook
from openpyxl.styles import Font

LOGGER = logging.getLogger("extract_html_procurement")

ENCODINGS = ("utf-8", "gbk", "gb18030")

# ----------------------------------------------------------------------
# Field configuration: adjust aliases / stopwords / regex as needed.
# ----------------------------------------------------------------------
FIELD_DEFINITIONS: Sequence[Dict] = (
    dict(
        name="公告时间",
        aliases=("公告时间", "发布日期", "发布时间", "发布公告时间", "信息时间"),
        multi=False,
        stopwords=(),
        fallback=(r"(公告时间|发布日期|发布时间|信息时间)[:：]?\s*(\d{4}[年/\-]\d{1,2}[月/\-]\d{1,2})",),
    ),
    dict(
        name="项目名称",
        aliases=("项目名称", "采购项目名称", "项目名称"),
        multi=False,
        stopwords=("采购单位", "采购人名称", "公告时间", "供应商名称", "中标金额"),
        fallback=(r"(项目名称|采购项目名称)[:：]?\s*([^\n\r]{2,120})",),
    ),
    dict(
        name="采购单位名称",
        aliases=("采购人名称", "采购单位"),
        multi=False,
        stopwords=(
            "采购单位地址",
            "采购人地址",
            "供应商名称",
            "中标金额",
            "成交金额",
            "采购人联系人",
            "采购人联系电话",
            "采购单位联系方式",
            "采购人联系方式",
            "联系方式",
            "联系电话",
            "项目联系人",
            "项目联系电话",
            "遴选专家名单",
            "行政区域",
        ),
        fallback=(r"(采购人名称|采购单位)[:：]?\s*([^\n\r]{2,120})",),
    ),
    dict(
        name="采购单位地址",
        aliases=("采购人地址", "采购单位地址"),
        multi=False,
        stopwords=(
            "采购单位名称",
            "供应商名称",
            "中标金额",
            "成交金额",
            "采购单位联系方式",
            "采购人联系方式",
            "代理机构名称",
            "代理机构地址",
            "代理机构",
            "附件",
        ),
        fallback=(r"(采购人地址|采购单位地址)[:：]?\s*([^\n\r]{2,160})",),
    ),
    dict(
        name="供应商名称",
        aliases=("供应商名称", "中标供应商", "成交供应商", "供应商", "中标人", "成交人"),
        multi=False,
        stopwords=(
            "供应商地址",
            "中标金额",
            "成交金额",
            "采购类别",
            "采购方式",
            "综合评分",
            "最终得分",
        ),
        fallback=(r"(供应商名称|中标供应商|成交供应商|中标人|成交人)[:：]?\s*([^\n\r]{2,160})",),
    ),
    dict(
        name="供应商地址",
        aliases=("供应商地址", "中标供应商地址", "成交供应商地址"),
        multi=False,
        stopwords=("供应商名称", "中标金额", "成交金额", "采购类别"),
        fallback=(r"(供应商地址|中标供应商地址|成交供应商地址)[:：]?\s*([^\n\r]{2,200})",),
    ),
    dict(
        name="中标金额",
        aliases=("中标金额", "中标（成交）金额", "成交金额", "合同金额"),
        multi=False,
        stopwords=(),
        fallback=(r"(中标金额|成交金额|合同金额)[:：]?\s*([^\n\r]{1,80})",),
    ),
    dict(
        name="采购类别",
        aliases=("采购类别", "采购类型", "采购方式", "项目类别", "品目"),
        multi=False,  # 改为单一值
        stopwords=(),
        fallback=(r"(采购类别|采购方式|项目类别|品目)[:：]?\s*([^\n\r]{1,80})",),
    ),
    dict(
        name="采购标的",
        aliases=("采购标的", "标的名称", "采购内容", "标的"),
        multi=False,
        stopwords=("品牌", "规格型号", "数量", "总价"),
        fallback=(r"(采购标的|标的名称|采购内容)[:：]?\s*([^\n\r]{2,120})",),
    ),
)

FIELD_ORDER: Sequence[str] = tuple(item["name"] for item in FIELD_DEFINITIONS)
MULTI_VALUE_FIELDS = {item["name"] for item in FIELD_DEFINITIONS if item["multi"]}
FIELD_KEYWORDS: Dict[str, Sequence[str]] = {item["name"]: item["aliases"] for item in FIELD_DEFINITIONS}
FIELD_STOPWORDS: Dict[str, Sequence[str]] = {item["name"]: item["stopwords"] for item in FIELD_DEFINITIONS}
FULL_TEXT_PATTERNS: Dict[str, Sequence[re.Pattern]] = {
    item["name"]: tuple(re.compile(pattern) for pattern in item["fallback"]) if item["fallback"] else ()
    for item in FIELD_DEFINITIONS
}

SUPPLIER_SCORE_TOKENS = (
    "公司",
    "有限",
    "集团",
    "中心",
    "医院",
    "大学",
    "学院",
    "学校",
    "研究",
    "科技",
    "股份",
    "合作社",
    "政府",
    "委员会",
    "事务所",
    "厂",
    "站",
    "局",
)
ADDRESS_SCORE_TOKENS = ("省", "市", "区", "县", "镇", "乡", "街", "路", "道", "大道", "村", "楼", "栋", "层", "单元", "室", "号")
LOCATION_SUFFIXES = ("省", "市", "区", "县", "镇", "乡", "街", "道", "大道", "办", "园")
ADMIN_LOCATION_SUFFIXES = ("省", "市", "区", "县", "镇", "乡")

CATEGORY_KEYWORDS = {
    "货物": ("货物", "设备", "物资", "用品", "耗材"),
    "服务": ("服务", "咨询", "保障", "培训", "运营", "维护"),
    "工程": ("工程", "施工", "改造", "维修", "建设"),
}

BLOCK_TAGS = {
    "p",
    "li",
    "span",
    "strong",
    "em",
    "b",
    "div",
    "td",
    "th",
    "dd",
    "dt",
    "h1",
    "h2",
    "h3",
    "h4",
    "h5",
    "h6",
}

TABLE_RANK_HEADERS = {
    "得分排名",
    "得分排位",
    "得分排名情况",
    "排名",
    "综合排名",
}
TABLE_RANK_ACCEPT = {"1", "第一名", "第一", "冠军"}


@dataclass
class FieldValue:
    multi: bool
    entries: List[Tuple[int, str]] = field(default_factory=list)

    def add(self, value: str, order: int) -> None:
        if not value:
            return
        cleaned = value.strip()
        if not cleaned:
            return
        if self.entries and not self.multi:
            return
        if self.multi and cleaned in {item[1] for item in self.entries}:
            return
        self.entries.append((order, cleaned))

    def get(self) -> str:
        if not self.entries:
            return ""
        if self.multi:
            ordered = [value for _, value in sorted(self.entries, key=lambda item: item[0])]
            unique: List[str] = []
            for value in ordered:
                if value not in unique:
                    unique.append(value)
            return "|".join(unique)
        return self.entries[0][1]


def setup_logging(verbose: bool = False) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S")


def read_html_text(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ENCODINGS:
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def normalize_whitespace(text: str) -> str:
    if not text:
        return ""
    return re.sub(r"\s+", " ", text.replace("\u3000", " ").replace("\xa0", " ")).strip()


def normalize_label_text(text: str) -> str:
    if not text:
        return ""
    cleaned = normalize_whitespace(text)
    cleaned = cleaned.replace("：", "").replace(":", "")
    cleaned = cleaned.replace("（", "").replace("）", "")
    cleaned = re.sub(r"^[一二三四五六七八九十\d]{1,3}[、.\-]", "", cleaned)
    return re.sub(r"\s+", "", cleaned).lower()


def normalize_text_for_regex(text: str) -> str:
    if not text:
        return ""
    text = text.replace("\u3000", " ").replace("\xa0", " ")
    text = text.replace("：", ":")
    return re.sub(r"\s+", " ", text).strip()


def cleanup_field_value(field: str, value: str) -> str:
    strip_chars = " 、。，；:"
    cleaned = value.strip(strip_chars)
    
    # 对于采购单位地址，先检查是否包含"代理机构"等关键词
    if field == "采购单位地址":
        # 如果包含代理机构相关关键词，直接过滤掉
        if re.search(r"(代理机构|招标代理|采购代理|中介机构)", cleaned):
            return ""
    
    # 对于项目名称和采购标的，不使用停止词过滤和冒号分割
    if field not in ("项目名称", "采购标的"):
        for stopword in FIELD_STOPWORDS.get(field, ()):
            idx = cleaned.find(stopword)
            if idx != -1:
                cleaned = cleaned[:idx]
    
    # 对于项目名称和采购标的，不按冒号分割
    if field not in ("项目名称", "采购标的"):
        for sep in ("：", ":"):
            if sep in cleaned:
                left, right = cleaned.split(sep, 1)
                if left.strip() and right.strip():
                    cleaned = left
                    break
    if field == "供应商名称":
        if any(token in cleaned for token in ("比例", "权重", "得分", "%")):
            if not re.search(r"(公司|有限|集团|中心|医院|大学|学院|学校|研究|科技|股份|合作社|政府|委员会|事务所|厂|站|局)", cleaned):
                cleaned = ""
    if field in {"供应商地址", "采购单位地址"}:
        for separator in ("；", ";"):
            idx = cleaned.find(separator)
            if idx != -1:
                cleaned = cleaned[:idx]
                break
    if field == "采购单位名称":
        idx = cleaned.find("联系方式")
        if idx != -1:
            cleaned = cleaned[:idx]
        # 过滤掉电话号码（包含数字且以特定模式开头）
        if re.match(r"^[\d\-\s]+$", cleaned) or re.match(r"^0\d{2,3}[-\s]?\d{7,8}$", cleaned):
            cleaned = ""
        # 过滤掉纯数字或看起来像电话号码的内容
        if cleaned and len(cleaned) < 20 and re.search(r"\d{7,}", cleaned) and not re.search(r"[\u4e00-\u9fa5]{2,}", cleaned):
            cleaned = ""
    
    # 注意：项目名称保留完整内容，不做额外清理
    
    return cleaned.strip(strip_chars)


def extract_location_tokens(text: str) -> Set[str]:
    tokens: Set[str] = set()
    if not text:
        return tokens
    for suffix in LOCATION_SUFFIXES:
        pattern = rf"[^\s\d]{{1,6}}{suffix}"
        for match in re.finditer(pattern, text):
            token = match.group(0)
            suffix_len = len(suffix)
            prefix_len = max(1, min(2, len(token) - suffix_len))
            trimmed = token[-(suffix_len + prefix_len):]
            tokens.add(trimmed)
    return tokens


def score_field_value(field: Optional[str], value: str, reference: str = "") -> int:
    if not value:
        return 0
    cleaned = value.strip()
    if not cleaned:
        return 0
    score = len(cleaned)
    
    # 对于项目名称，更长更完整的标题得分更高（包含项目编号和公告类型的版本）
    if field == "项目名称":
        # 如果包含"项目编号"，给予额外加分
        if "项目编号" in cleaned or "编号" in cleaned:
            score += 50
        # 如果包含公告类型，给予额外加分
        if re.search(r"(中标|成交|结果)公告", cleaned):
            score += 30
    
    if re.search(r"(详见|另行|暂未|待定|--|见公告|见附件|无)", cleaned):
        score -= 20
    
    # 对于采购标的字段,如果包含版本号、型号等特征,降低分数
    # 因为这更可能是规格型号而不是标的名称
    if field == "采购标的":
        # 检测是否包含版本号、型号等特征
        if re.search(r'(V\d+\.\d+|版本|型号|简称[:：]|规格[:：]|\[|\]|（.*?版.*?）)', cleaned):
            score -= 100  # 大幅降低分数,避免规格型号覆盖标的名称
    
    if field in {"供应商地址", "采购单位地址"}:
        for token in ADDRESS_SCORE_TOKENS:
            if token in cleaned:
                score += 3
        if re.search(r"\d", cleaned):
            score += 3
    if field in {"供应商地址", "采购单位地址"} and reference:
        reference_tokens = extract_location_tokens(reference)
        if reference_tokens:
            candidate_tokens = extract_location_tokens(cleaned)
            matches = candidate_tokens & reference_tokens
            if matches:
                bonus = 20 if field == "采购单位地址" else 10
                score += bonus * len(matches)
            if field == "采购单位地址" and candidate_tokens:
                mismatches = {
                    token
                    for token in candidate_tokens
                    if token not in reference_tokens and token[-1] in ADMIN_LOCATION_SUFFIXES
                }
                if mismatches:
                    score -= 6 * len(mismatches)
    if field == "供应商名称":
        for token in SUPPLIER_SCORE_TOKENS:
            if token in cleaned:
                score += 4
    if field == "采购单位名称":
        for token in ("局", "委", "院", "中心", "办", "公司", "集团", "学校", "医院", "大学", "政府", "管理"):
            if token in cleaned:
                score += 2
    return score


def normalize_date_text(value: str) -> Optional[str]:
    if not value:
        return None
    text = value.strip()
    if not text:
        return None
    text = text.replace("年", "-").replace("月", "-").replace("日", "")
    text = text.replace("/", "-").replace(".", "-")
    text = re.sub(r"-+", "-", text)
    match = re.search(r"\d{4}-\d{1,2}-\d{1,2}", text)
    candidate = match.group(0) if match else text
    try:
        dt = date_parser.parse(candidate, dayfirst=False, yearfirst=True, fuzzy=True)
        # 统一格式：个位数补0
        return f"{dt.year:04d}年{dt.month:02d}月{dt.day:02d}日"
    except (ValueError, OverflowError):
        return None


def normalize_amounts(raw: str) -> List[str]:
    if not raw:
        return []
    text = raw.replace(",", "").replace(" ", "")
    text = text.replace("人民币", "").replace("万元整", "万元")
    patterns = re.finditer(r"([\d,.]+)(万元|元)?", text)
    results: List[str] = []
    for match in patterns:
        number, unit = match.groups()
        if not number:
            continue
        number = number.replace(",", "")
        try:
            amount = Decimal(number)
        except InvalidOperation:
            continue
        if unit == "万元":
            amount *= Decimal("10000")
        normalized = f"{amount:.2f}"
        if normalized not in results:
            results.append(normalized)
    return results


def extract_categories(raw: str) -> List[str]:
    if not raw:
        return []
    # 移除空格，处理"货物"和"类"被分隔的情况
    text = re.sub(r'\s+', '', raw).lower()
    detected: List[str] = []
    # 先检查明确的类别标识（如"货物类"、"服务类"、"工程类"）
    if "货物类" in text:
        return ["货物"]
    if "服务类" in text:
        return ["服务"]
    if "工程类" in text:
        return ["工程"]
    
    # 如果没有找到明确的类别标识，再使用关键词匹配
    text_with_space = normalize_whitespace(raw).lower()
    for category, keywords in CATEGORY_KEYWORDS.items():
        if any(keyword.lower() in text_with_space for keyword in keywords):
            detected.append(category)
    return detected


def label_to_field(label: str) -> Optional[str]:
    normalized = normalize_label_text(label)
    if not normalized:
        return None
    best_field: Optional[str] = None
    best_score = 0
    for field, aliases in FIELD_KEYWORDS.items():
        for alias in aliases:
            alias_norm = normalize_label_text(alias)
            if not alias_norm:
                continue
            if normalized == alias_norm:
                score = len(alias_norm) + 100
            elif alias_norm in normalized:
                score = len(alias_norm)
            elif normalized in alias_norm:
                # 只有当标签也包含别名时才匹配（太宽松，容易误匹配）
                # 例如："地址" in "采购单位地址" 会匹配，但我们不希望单独的"地址"被映射
                # 解决方案：对于地址字段，要求必须包含"采购"或"供应商"等前缀
                if field in {"采购单位地址", "供应商地址"}:
                    # 检查原始标签是否包含必要的前缀
                    if field == "采购单位地址":
                        if not ("采购" in label or "采购人" in label):
                            # 不匹配单独的"地址"、"地 址"等
                            continue
                    elif field == "供应商地址":
                        if not ("供应商" in label or "中标供应商" in label or "成交供应商" in label):
                            continue
                score = len(normalized)
            else:
                continue
            if score > best_score:
                best_field = field
                best_score = score
    return best_field


class ProcurementExtractor:
    def __init__(self, soup: BeautifulSoup):
        self.soup = soup
        self.fields = {name: FieldValue(multi=name in MULTI_VALUE_FIELDS) for name in FIELD_ORDER}
        self.counter = 0
        self.full_text = normalize_text_for_regex(soup.get_text("\n"))
        self._processed_tables: Set[int] = set()

    def _next_order(self) -> int:
        self.counter += 1
        return self.counter

    def add_field(self, field: str, raw_value: str) -> None:
        if field not in self.fields or not raw_value:
            return
        field_store = self.fields[field]
        
        if field == "公告时间":
            normalized = normalize_date_text(raw_value)
            values = [normalized] if normalized else []
        elif field == "中标金额":
            values = normalize_amounts(raw_value)
        elif field == "采购类别":
            values = extract_categories(raw_value)
        else:
            cleaned = cleanup_field_value(field, normalize_whitespace(raw_value))
            values = [cleaned] if cleaned else []
        reference = ""
        if field == "采购单位地址":
            name_entries = self.fields["采购单位名称"].entries
            if name_entries:
                reference = name_entries[0][1]
        elif field == "供应商地址":
            supplier_entries = self.fields["供应商名称"].entries
            if supplier_entries:
                reference = supplier_entries[0][1]
        for value in values:
            if field_store.entries and not field_store.multi:
                existing_value = field_store.entries[0][1]
                existing_score = score_field_value(field, existing_value, reference)
                candidate_score = score_field_value(field, value, reference)
                
                if candidate_score > existing_score:
                    field_store.entries[0] = (self._next_order(), value.strip())
                continue
            field_store.add(value, self._next_order())

    def _extract_meta_nodes(self) -> None:
        meta_title = self.soup.find("meta", attrs={"name": re.compile(r"^ArticleTitle$", re.I)})
        if meta_title:
            value = meta_title.get("content") or meta_title.get("value")
            if value:
                self.add_field("项目名称", value)
        meta_date = self.soup.find("meta", attrs={"name": re.compile(r"^PubDate$", re.I)})
        if meta_date:
            value = meta_date.get("content") or meta_date.get("value")
            if value:
                self.add_field("公告时间", value)

    def extract(self) -> Dict[str, str]:
        self._extract_meta_nodes()
        body = self.soup.body or self.soup
        if body:
            self._walk_dom(body)
        self._apply_full_text_patterns()
        result = {field: value.get() for field, value in self.fields.items()}
        
        # 特殊处理采购类别：优先从全文中查找"货物类"、"服务类"、"工程类"
        # 移除空格处理被分隔的情况（如"货物"和"类"在不同的span中）
        full_text_no_space = re.sub(r'\s+', '', self.full_text)
        if not result["采购类别"] or result["采购类别"] == "其他":
            # 在全文中查找明确的类别标识（移除空格版本）
            if "货物类" in full_text_no_space:
                result["采购类别"] = "货物"
            elif "服务类" in full_text_no_space:
                result["采购类别"] = "服务"
            elif "工程类" in full_text_no_space:
                result["采购类别"] = "工程"
            else:
                # 如果没有找到明确的类别，保持原值或设为"其他"
                if not result["采购类别"]:
                    result["采购类别"] = "其他"
        
        return result

    def _walk_dom(self, node: Tag) -> None:
        for element in node.children:
            if isinstance(element, NavigableString):
                continue
            if not isinstance(element, Tag):
                continue
            if element.name == "table":
                table_id = id(element)
                if table_id not in self._processed_tables:
                    self._processed_tables.add(table_id)
                    self._extract_table(element)
                continue
            if element.find_parent("table") is not None:
                continue
            if element.name in BLOCK_TAGS:
                text = element.get_text(separator="\n", strip=True)
                if text:
                    self._process_text_block(text)
            self._walk_dom(element)

    def _process_text_block(self, text: str) -> None:
        lines = [line.strip() for line in text.replace("\r", "\n").split("\n") if line.strip()]
        if not lines:
            return
        merged: List[str] = []
        skip_next = False
        for idx, line in enumerate(lines):
            if skip_next:
                skip_next = False
                continue
            if re.search(r"[；;]$", line) and idx + 1 < len(lines):
                merged.append(f"{line} {lines[idx + 1]}")
                skip_next = True
            else:
                merged.append(line)
        
        # 标记是否在代理机构相关区域
        in_agency_section = False
        
        for segment in merged:
            # 检测是否进入代理机构信息区域（更宽松的匹配）
            if re.search(r"(代理机构|采购代理|招标代理|中介机构)", segment):
                in_agency_section = True
            # 如果再次出现采购人信息或项目联系，则退出代理机构区域
            if re.search(r"(采购人信息|采购单位信息|项目联系|其他补充|附件)", segment) and "代理" not in segment:
                in_agency_section = False
            
            for candidate in re.split(r"[；;]", segment):
                candidate = candidate.strip()
                if not candidate:
                    continue
                match = re.match(r"(.{1,40}?)[：:]\s*(.+)", candidate)
                label = None
                value = None
                if match:
                    label, value = match.group(1), match.group(2)
                else:
                    alt = re.match(r"(.{1,40}?)(供应商名称|供应商地址|采购单位名称|采购单位地址)[：:\s]+(.+)", candidate)
                    if alt:
                        prefix, label, value = alt.groups()
                        if prefix.strip().isdigit():
                            # keep label and value
                            pass
                        else:
                            label = None
                if label and value:
                    field = label_to_field(label)
                    if not field:
                        extra = re.match(r"(.{1,40}?)[：:\s]+(.+)", value)
                        if extra:
                            alt_label, alt_value = extra.groups()
                            field = label_to_field(alt_label)
                            if field:
                                value = alt_value
                    if field:
                        # 如果是采购单位地址字段，且在代理机构区域，跳过
                        if field == "采购单位地址" and in_agency_section:
                            continue
                        
                        self.add_field(field, value)

    def _extract_table(self, table: Tag) -> None:
        # 重写表格解析逻辑,正确处理rowspan/colspan
        trs = table.find_all("tr")
        # 构建完整的表格矩阵,处理rowspan和colspan
        matrix: List[List[Optional[str]]] = []
        cell_sources: List[List[Optional[int]]] = []  # 记录每个单元格来自哪一行(用于调试)
        
        for row_idx, tr in enumerate(trs):
            cells = tr.find_all(["th", "td"])
            if row_idx >= len(matrix):
                matrix.append([])
                cell_sources.append([])
            
            col_idx = 0
            for cell in cells:
                # 跳过被前面的rowspan占用的列
                while col_idx < len(matrix[row_idx]) and matrix[row_idx][col_idx] is not None:
                    col_idx += 1
                
                # 获取单元格的文本内容
                cell_text = normalize_whitespace(cell.get_text(separator=" ", strip=True))
                
                # 获取rowspan和colspan属性
                rowspan = int(cell.get("rowspan", 1))
                colspan = int(cell.get("colspan", 1))
                
                # 填充当前单元格及其跨越的区域
                for r in range(rowspan):
                    target_row = row_idx + r
                    # 确保目标行存在
                    while len(matrix) <= target_row:
                        matrix.append([])
                        cell_sources.append([])
                    # 确保目标行足够长
                    while len(matrix[target_row]) <= col_idx + colspan - 1:
                        matrix[target_row].append(None)
                        cell_sources[target_row].append(None)
                    
                    # 填充跨越的列
                    for c in range(colspan):
                        matrix[target_row][col_idx + c] = cell_text
                        cell_sources[target_row][col_idx + c] = row_idx
                
                col_idx += colspan
        
        # 现在matrix包含了正确对齐的表格数据
        rows = matrix
        header_map: Dict[int, str] = {}
        ranking_idx: Optional[int] = None
        
        # 跟踪采购单位和代理机构的名称,用于判断地址归属
        last_name_field = None  # 最近遇到的名称字段类型
        
        for row_idx, row in enumerate(rows):
            # 过滤掉None值
            row = [cell if cell is not None else "" for cell in row]
            candidate_map = {idx: label_to_field(cell) for idx, cell in enumerate(row)}
            candidate_map = {idx: field for idx, field in candidate_map.items() if field}
            if not header_map and candidate_map:
                if len(candidate_map) >= 2 or len(row) > 2:
                    header_map = candidate_map
                    for idx, cell in enumerate(row):
                        if normalize_label_text(cell) in {normalize_label_text(x) for x in TABLE_RANK_HEADERS}:
                            ranking_idx = idx
                            break
                    continue
            if header_map:
                if ranking_idx is not None and ranking_idx < len(row):
                    rank_norm = normalize_label_text(row[ranking_idx])
                    if rank_norm not in {normalize_label_text(x) for x in TABLE_RANK_ACCEPT}:
                        continue
                for idx, field in header_map.items():
                    if idx < len(row):
                        value = row[idx]
                        
                        if label_to_field(value):
                            continue
                        # 记录原始值,用于调试
                        original_value = value
                        # 特殊处理:"采购标的"字段的智能识别
                        if field == "采购标的":
                            # 检查是否包含通用模板描述（这些通常不是真正的标的名称）
                            generic_patterns = [
                                r'按.*[《<].*招标.*文件.*[》>].*要求.*执行',
                                r'按.*[《<].*投标.*文件.*[》>].*执行',
                                r'详见.*文件',
                                r'详见.*公告',
                                r'见.*附件'
                            ]
                            is_generic = any(re.search(pattern, value) for pattern in generic_patterns)
                            
                            if is_generic or (len(value.strip()) <= 4 and not re.search(r'(系统|设备|服务|工程|项目|采购|建设)', value)):
                                # 尝试查找同行中更合适的值（优先选择包含项目关键词的列）
                                best_alt_value = None
                                best_score = 0
                                
                                for alt_idx in range(len(row)):
                                    if alt_idx != idx and alt_idx not in header_map:
                                        alt_value = row[alt_idx].strip()
                                        if not alt_value or alt_value == value:
                                            continue
                                        
                                        # 计算候选值的得分
                                        score = len(alt_value)
                                        if re.search(r'(系统|设备|服务|工程|项目|采购|建设|平台|软件|硬件)', alt_value):
                                            score += 20
                                        # 排除看起来像通用描述的值
                                        if any(re.search(p, alt_value) for p in generic_patterns):
                                            score = 0
                                        
                                        if score > best_score:
                                            best_score = score
                                            best_alt_value = alt_value
                                
                                if best_alt_value:
                                    value = best_alt_value
                        
                        self.add_field(field, value)
            if len(row) >= 2:
                if ranking_idx is not None and ranking_idx < len(row):
                    rank_norm = normalize_label_text(row[ranking_idx])
                    if rank_norm not in {normalize_label_text(x) for x in TABLE_RANK_ACCEPT}:
                        continue
                for idx in range(0, len(row) - 1, 2):
                    label_text = row[idx]
                    value_text = row[idx + 1]
                    field = label_to_field(label_text)
                    
                    if field:
                        # 检查标签文本本身是否包含"代理"关键词
                        label_has_agency = "代理" in label_text
                        
                        # 跟踪当前处理的是采购单位还是代理机构
                        if field == "采购单位名称" and not label_has_agency:
                            last_name_field = "purchaser"
                        elif label_has_agency or "代理机构" in label_text:
                            last_name_field = "agency"
                        
                        # 对于地址字段的特殊处理
                        if field == "采购单位地址":
                            # 1. 如果标签本身包含"代理"，跳过
                            if label_has_agency:
                                continue
                            # 2. 如果最近处理的是代理机构名称，跳过
                            if last_name_field == "agency":
                                continue
                        
                        self.add_field(field, value_text)

    def _apply_full_text_patterns(self) -> None:
        for field, patterns in FULL_TEXT_PATTERNS.items():
            if self.fields[field].entries:
                continue
            for pattern in patterns:
                for match in pattern.finditer(self.full_text):
                    value = match.group(2) if match.lastindex and match.lastindex >= 2 else match.group(1)
                    self.add_field(field, value)
                    if self.fields[field].entries and not self.fields[field].multi:
                        break
                if self.fields[field].entries and not self.fields[field].multi:
                    break


def build_record(path: Path) -> Dict[str, str]:
    soup = BeautifulSoup(read_html_text(path), "lxml")
    extractor = ProcurementExtractor(soup)
    return extractor.extract()


def extract_directory(input_dir: Path) -> Tuple[List[Dict[str, str]], Dict[str, int]]:
    records: List[Dict[str, str]] = []
    stats = {"total": 0, "success": 0, "failed": 0}
    for html_file in sorted(input_dir.rglob("*.htm*")):
        stats["total"] += 1
        try:
            record = build_record(html_file)
        except Exception as exc:  # pylint: disable=broad-except
            LOGGER.exception("Failed to process %s: %s", html_file, exc)
            stats["failed"] += 1
            continue
        records.append(record)
        stats["success"] += 1
    return records, stats


def write_outputs(records: Sequence[Dict[str, str]], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / "extracted.csv"
    jsonl_path = output_dir / "extracted.jsonl"
    xlsx_path = output_dir / "extracted.xlsx"

    # 写入CSV文件
    with csv_path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=list(FIELD_ORDER))
        writer.writeheader()
        for record in records:
            # 为"公告时间"字段使用Excel公式格式,防止Excel自动格式化去掉前导0
            # 格式: ="2021年05月24日",Excel会显示为文本且不显示公式符号
            row_data = {}
            for field in FIELD_ORDER:
                value = record.get(field, "")
                if field == "公告时间" and value:
                    # 使用Excel文本公式格式,保留前导0且不显示公式
                    value = f'="{value}"'
                row_data[field] = value
            writer.writerow(row_data)

    # 写入JSONL文件
    with jsonl_path.open("w", encoding="utf-8") as jsonl_file:
        for record in records:
            jsonl_file.write(json.dumps(record, ensure_ascii=False) + "\n")
    
    # 写入XLSX文件(Excel原生格式,完美支持文本格式的日期)
    wb = Workbook()
    ws = wb.active
    ws.title = "采购信息"
    
    # 写入表头(加粗)
    header_font = Font(bold=True)
    for col_idx, field in enumerate(FIELD_ORDER, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
    
    # 写入数据行
    for row_idx, record in enumerate(records, start=2):
        for col_idx, field in enumerate(FIELD_ORDER, start=1):
            value = record.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # 将"公告时间"列设置为文本格式,保留前导0
            if field == "公告时间":
                cell.number_format = '@'  # '@'表示文本格式
    
    # 自动调整列宽
    for col_idx, field in enumerate(FIELD_ORDER, start=1):
        # 根据字段名长度和内容估算列宽
        max_length = len(field)
        for row_idx in range(2, min(len(records) + 2, 100)):  # 只检查前100行
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        # 设置列宽(Excel列宽单位与字符数不完全一致,需要微调)
        adjusted_width = min(max_length + 2, 60)  # 最大60个字符宽度
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width
    
    wb.save(xlsx_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="批量提取 HTML 公告中的采购信息")
    parser.add_argument("--input", type=Path, default=Path("./data"), help="输入目录 (默认: ./data)")
    parser.add_argument("--output", type=Path, default=Path("./result"), help="输出目录 (默认: ./result)")
    parser.add_argument("--verbose", action="store_true", help="输出详细日志")
    args = parser.parse_args()

    setup_logging(verbose=args.verbose)

    records, stats = extract_directory(args.input)
    LOGGER.info("Processed %d files: %d succeeded, %d failed", stats["total"], stats["success"], stats["failed"])

    if records:
        write_outputs(records, args.output)
        LOGGER.info("Results written to %s", args.output)
    else:
        LOGGER.warning("No records extracted; outputs not written.")


if __name__ == "__main__":
    main()
