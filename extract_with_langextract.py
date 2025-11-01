#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""使用 LangExtract + DeepSeek 批量抽取政府采购公告字段。"""

from __future__ import annotations

import csv,os
import json
import logging
import sys
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from datetime import datetime
import time
from openpyxl import Workbook
from openpyxl.styles import Font

try:
    from langextract.providers.openai import OpenAILanguageModel
except ImportError:  # pragma: no cover - handled lazily in configure_model
    OpenAILanguageModel = None  # type: ignore[assignment]

from html_preprocess import (
    DEFAULT_ENCODINGS,
    DEFAULT_PATTERNS,
    collect_html_paths,
    convert_html_file,
)

try:
    from absl import logging as absl_logging
except ImportError:  # pragma: no cover - absl 非必需依赖
    absl_logging = None  # type: ignore[assignment]

LOGGER = logging.getLogger("langextract_pipeline")

MODEL_ID = "deepseek-chat"
API_KEY = os.environ.get('DEEPSEEK_API_KEY')
BASE_URL = "https://api.deepseek.com"
MODEL_TEMPERATURE = 0  # 低温度让输出更加稳定

FIELD_NAMES: Sequence[str] = (
    "公告时间",
    "项目名称",
    "采购单位名称",
    "采购单位地址",
    "供应商名称",
    "供应商地址",
    "中标金额",
    "采购类别",
    "采购标的",
)

OUTPUT_FIELDS: Sequence[str] = (*FIELD_NAMES, "来源文件")

EXAMPLE_LINES: Sequence[Tuple[str, str]] = (
    ("公告时间", "2024年05月02日"),
    ("项目名称", "城市道路维护项目（中标公告）"),
    ("采购单位名称", "某市城市管理局"),
    ("采购单位地址", "某市新区建设大道88号"),
    ("供应商名称", "某某建设有限公司"),
    ("供应商地址", "某市高新区产业园一区"),
    ("中标金额", "1980000元"),
    ("采购类别", "服务类"),
    ("采购标的", "城市道路日常维护服务"),
)

# 屏蔽 LangExtract 的部分告警，避免干扰输出
warnings.filterwarnings(
    "ignore",
    message="'use_schema_constraints' is ignored when 'model' is provided.*",
)
if absl_logging is not None:
    absl_logging.set_verbosity(absl_logging.ERROR)


@dataclass
class PipelineConfig:
    input_dir: Path = Path("./data")
    output_dir: Path = field(
        default_factory=lambda: Path("./result") / datetime.now().strftime("%Y%m%d_%H%M%S")
    )
    extraction_passes: int = 1
    max_char_buffer: int = 2400
    max_workers: int = 5
    use_schema_constraints: bool = True
    fence_output: Optional[bool] = None
    resolver_params: Optional[Dict[str, Any]] = None
    log_level: int = logging.INFO
    patterns: Sequence[str] = DEFAULT_PATTERNS
    encodings: Sequence[str] = DEFAULT_ENCODINGS
    separator: str = "\n"
    drop_empty_lines: bool = True
    table_cell_sep: str = " | "
    fail_silently: bool = False
    default_model_id: str = "gemini-2.5-flash"
    use_custom_model: bool = True


CONFIG = PipelineConfig()


@dataclass
class PipelineStats:
    total: int = 0
    success: int = 0
    failed: int = 0


def configure_model(lx_module: Any, config: PipelineConfig) -> Optional[Any]:
    """返回 DeepSeek 模型包装。"""
    if not config.use_custom_model:
        return None
    if OpenAILanguageModel is None:
        raise RuntimeError("未安装 langextract，请先执行：pip install langextract")
    return OpenAILanguageModel(
        model_id=MODEL_ID,
        api_key=API_KEY,
        base_url=BASE_URL,
        temperature=MODEL_TEMPERATURE,
        format_type=lx_module.data.FormatType.JSON,
    )


def build_prompt(lx_module: Any) -> Tuple[str, Sequence[Any]]:
    """构造 LangExtract 所需的提示和 few-shot 示例。"""
    field_list = "、".join(FIELD_NAMES)
    prompt = (
        "请从政府采购公告中提取以下字段："
        f"{field_list}。"
        "\n- 公告时间：公告发布日期时间（YYYY年MM月DD日 HH:MM格式），优先提取“公告时间 |”后的完整时间"
        "\n- 项目名称：完整的采购项目名称（优先从“采购项目名称 |”或“二、项目名称：”提取）"
        "\n- 采购单位名称：发起采购的机构全称（优先从“采购单位 |”提取）"
        "\n- 采购单位地址：采购单位完整地址（优先从“采购单位地址 |”提取，需包含门牌号）"
        "\n- 供应商名称：中标供应商全称（从“采购结果”的表格提取），多个请用；隔开"
        "\n- 供应商地址：中标供应商完整地址（从采购结果表格中对应行提取），多个请用；隔开"
        "\n- 中标金额：总中标金额（从“总中标金额 |”提取，仅包含数字和单位即可）"
        "\n- 采购类别：品目类别（如“服务类”“货物类”以及“工程类”，从主要标的信息部分提取，多个请用；隔开）"
        "\n- 采购标的：具体采购内容（主要从“采购标的”列提取，从主要标的信息部分提取，如“农畜产品批发服务”，多个请用；隔开）"
        "\n每个字段最多输出一条，且必须来自公告原文的连续片段；缺失则留空。"
    )
    example_text = "\n".join(f"{key}：{value}" for key, value in EXAMPLE_LINES)
    example_extractions = [
        lx_module.data.Extraction(extraction_class=key, extraction_text=value)
        for key, value in EXAMPLE_LINES
    ]
    example = lx_module.data.ExampleData(text=example_text, extractions=example_extractions)
    return prompt, [example]

class ProgressBar:
    """在终端打印简单的进度条。"""

    def __init__(self, total: int, width: int = 30) -> None:
        self.total = total
        self.width = width
        self._finished = False

    def update(self, current: int, path: Optional[Path] = None) -> None:
        if self.total <= 0:
            return
        ratio = max(0.0, min(1.0, current / self.total))
        filled = int(self.width * ratio)
        bar = "#" * filled + "-" * (self.width - filled)
        suffix = f" {path.name}" if path is not None else ""
        sys.stdout.write(
            f"\r处理进度 [{bar}] {current}/{self.total} ({ratio * 100:5.1f}%){suffix}"
        )
        sys.stdout.flush()
        if current >= self.total:
            self.finish()

    def finish(self) -> None:
        if self.total <= 0 or self._finished:
            return
        sys.stdout.write("\n")
        sys.stdout.flush()
        self._finished = True


class LangExtractPipeline:
    """负责批量执行 LangExtract 并写入结构化结果。"""

    def __init__(self, lx_module: Any, config: PipelineConfig):
        self.lx = lx_module
        self.config = config
        self.model = configure_model(lx_module, config)
        self.prompt, self.examples = build_prompt(lx_module)

    def _build_extract_kwargs(self, text: str) -> Dict[str, Any]:
        kwargs: Dict[str, Any] = {
            # 待抽取的原文文本
            "text_or_documents": text,
            # 针对本任务定制的提示词
            "prompt_description": self.prompt,
            # few-shot 示例，帮助模型理解字段格式
            "examples": self.examples,
            # 允许多次抽取以提升召回
            "extraction_passes": self.config.extraction_passes,
            # 控制单次推理的字符上限
            "max_char_buffer": len(text),
            # 并发 worker 数，影响吞吐
            "max_workers": self.config.max_workers,
            # 是否让 LangExtract 自动生成 schema 约束
            "use_schema_constraints": self.config.use_schema_constraints,
            # 明确表示文本来自本地，不做网络抓取
            "fetch_urls": False,
        }
        if self.model is not None:
            kwargs["model"] = self.model
        else:
            kwargs["model_id"] = self.config.default_model_id
        if self.config.fence_output is not None:
            kwargs["fence_output"] = self.config.fence_output
        if self.config.resolver_params:
            kwargs["resolver_params"] = self.config.resolver_params
        kwargs["debug"] = False
        return kwargs

    def extract_record(self, text: str) -> Dict[str, str]:
        annotated = self.lx.extract(**self._build_extract_kwargs(text))
        self.lx.io.save_annotated_documents(
            [annotated],
            output_name="extraction_results_test.jsonl",
            output_dir="."
        )
        if hasattr(annotated, 'extractions') and annotated.extractions:
            print("\n" + "="*60)
            print("📊 提取详情:")
            print("="*60)
            for i, ext in enumerate(annotated.extractions, 1):
                print(f"\n{i}. {ext.extraction_class}")
                print(f"   文本: {ext.extraction_text}")
                print(f"   属性: {ext.attributes}")
            
        documents = annotated if isinstance(annotated, list) else [annotated]
        record = {field: "" for field in FIELD_NAMES}
        for document in documents:
            for extraction in getattr(document, "extractions", []) or []:
                field_name = extraction.extraction_class
                if field_name not in record or record[field_name]:
                    continue
                record[field_name] = (extraction.extraction_text or "").strip()
        return record

    def process_documents(self) -> Tuple[List[Dict[str, str]], PipelineStats]:
        stats = PipelineStats()
        records: List[Dict[str, str]] = []

        files = collect_html_paths(self.config.input_dir, patterns=self.config.patterns)
        stats.total = len(files)

        if stats.total == 0:
            LOGGER.info("未检测到可处理的 HTML 文件，提前退出。")
            return records, stats

        LOGGER.info("本次共有 %d 个文件待处理", stats.total)
        print(f"📂 本次检测到 {stats.total} 个 HTML 文件，开始执行抽取...")
        progress = ProgressBar(stats.total)

        for index, path in enumerate(files, start=1):
            try:
                text = convert_html_file(
                    path,
                    encodings=self.config.encodings,
                    separator=self.config.separator,
                    drop_empty_lines=self.config.drop_empty_lines,
                    table_cell_sep=self.config.table_cell_sep,
                )
            except Exception as exc:  # pylint: disable=broad-except
                stats.failed += 1
                if self.config.fail_silently:
                    LOGGER.warning("预处理失败：%s -> %s", path, exc)
                else:
                    LOGGER.exception("预处理失败：%s", path)
            else:
                try:
                    record = self.extract_record(text)
                except Exception:  # pylint: disable=broad-except
                    stats.failed += 1
                    LOGGER.exception("抽取失败：%s", path)
                else:
                    record["来源文件"] = str(path)
                    records.append(record)
                    stats.success += 1
            finally:
                progress.update(index, path)

        progress.finish()
        return records, stats


def write_outputs(records: Sequence[Dict[str, str]], output_dir: Path) -> None:
    """将抽取结果写入 CSV / JSONL / XLSX。"""
    output_dir.mkdir(parents=True, exist_ok=True)

    csv_path = output_dir / "extracted.csv"
    jsonl_path = output_dir / "extracted.jsonl"
    xlsx_path = output_dir / "extracted.xlsx"

    with csv_path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=list(OUTPUT_FIELDS), extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)

    with jsonl_path.open("w", encoding="utf-8") as jsonl_file:
        for record in records:
            jsonl_file.write(json.dumps(record, ensure_ascii=False) + "\n")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "采购信息"
    header_font = Font(bold=True)
    for col_idx, field in enumerate(OUTPUT_FIELDS, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=field)
        cell.font = header_font

    for row_idx, record in enumerate(records, start=2):
        for col_idx, field in enumerate(OUTPUT_FIELDS, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=record.get(field, ""))

    workbook.save(xlsx_path)


def setup_logging(level: int) -> None:
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        force=True,
    )
    LOGGER.setLevel(CONFIG.log_level)


def load_langextract() -> Any:
    try:
        import langextract as lx  # type: ignore[import-not-found]
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "当前环境未安装 langextract，请先运行：pip install langextract"
        ) from exc
    return lx


def run() -> None:
    setup_logging(CONFIG.log_level)
    lx_module = load_langextract()
    pipeline = LangExtractPipeline(lx_module, CONFIG)
    active_model_id = getattr(pipeline.model, "model_id", None) or CONFIG.default_model_id
    # 记录开始时间
    start_time = time.time()
    print(f"🚀 使用模型: {active_model_id}")
    records, stats = pipeline.process_documents()
    if records:
        write_outputs(records, CONFIG.output_dir)
        LOGGER.info(
            "处理完成：共 %d 个文件，成功 %d，失败 %d",
            stats.total,
            stats.success,
            stats.failed,
        )
        LOGGER.info("结果已保存至 %s", CONFIG.output_dir.resolve())
    else:
        LOGGER.warning("未成功抽取任何记录，请检查输入数据或模型配置。")
    # 记录结束时间
    end_time = time.time()
    elapsed_time = end_time - start_time
    LOGGER.info("处理耗时: %.2f 秒", elapsed_time)  

if __name__ == "__main__":
    run()
